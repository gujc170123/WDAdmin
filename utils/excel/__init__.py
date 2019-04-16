# -*- coding:utf-8 -*-
from __future__ import unicode_literals

import os

import time

import datetime
import xlrd
import xlwt
# import pandas
from utils.regular import RegularUtils


def str_check(str_obj):
    if type(str_obj) == int or type(str_obj) == long:
        str_obj = str(long(str_obj))
    elif type(str_obj) == float:
        str_obj = str(long(str_obj))
    return str_obj


try:
    with_openpyxl = True
    import openpyxl
except:
    with_openpyxl = False
    print "can not find openpyxl"

from WeiDuAdmin.settings import BASE_DIR


class ExcelUtils(object):

    def __init__(self):
        # self.wbk = xlwt.Workbook()
        self.wbk = openpyxl.Workbook()
        # writer = pandas.ExcelWriter(file, engine='xlsxwriter')

    def read_rows(self, excel_path, sheet_index=0):
        data = xlrd.open_workbook(excel_path)
        table = data.sheet_by_index(sheet_index)
        row_count = table.nrows
        for i in xrange(row_count):
            yield table.row_values(i)

    def create_excel(self, excel_name, titles, data, sheet_name="sheet1", force_save=True, parent_dir=None, sheet_index=None):
        # sheet = self.wbk.add_sheet(sheet_name, cell_overwrite_ok=True)
        sheet = self.wbk.create_sheet(sheet_name, index=sheet_index)
        self.write_title(sheet, titles)
        self.write_data(sheet, data)
        now = datetime.datetime.now()
        str_today = now.strftime("%Y-%m-%d")
        if parent_dir is None:
            parent_dir = os.path.join(BASE_DIR, "download", str_today)
        if not os.path.exists(parent_dir):
            os.mkdir(parent_dir)
        file_full_path = os.path.join(parent_dir, excel_name).encode("utf-8")
        if force_save:
            self.save_excel(file_full_path)
        return file_full_path

    def save_excel(self, file_full_path):
        self.wbk.save(file_full_path)
        return file_full_path

    def write_title(self, sheet_obj, titles):
        self.write_sheet_row(sheet_obj, titles, 0, True)

    def write_data(self, sheet_obj, data):
        for i, d in enumerate(data):
            self.write_sheet_row(sheet_obj, d, i+1)

    def write_sheet_row(self, sheet_obj, rowValueList, rowIndex, isBold=False):
        u"""往EXCEl单元格写内容，每次写一行sheet:页签名称；row：行内容列表；rowIndex：行索引;
        　　isBold:true:粗字段，false:普通字体"""
        if not with_openpyxl:
            i = 0
            style = xlwt.easyxf('font: bold 1', num_format_str='M/D/YY')
            # style = xlwt.easyxf('pattern: pattern solid, fore_color red;')
            for svalue in rowValueList:
                if isBold:
                    sheet_obj.write(rowIndex, i, svalue, style)
                else:
                    sheet_obj.write(rowIndex, i, svalue)
                i += 1
        else:
            sheet_obj.append(rowValueList)